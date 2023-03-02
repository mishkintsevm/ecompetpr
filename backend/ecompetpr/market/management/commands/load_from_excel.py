from django.core.management.base import BaseCommand

from ecompetpr.settings import DATA_DIR

from openpyxl import load_workbook

from market.models import Category, Product


class Command(BaseCommand):

    def handle(self, *args, **options):
        print('Clearing DB')
        Category.objects.all().delete()
        Product.objects.all().delete()

        print('Start importing from excel %s' % DATA_DIR)

        wb = load_workbook(DATA_DIR+'/price.xlsx')
        sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        print(sheet.max_row)
        for count in range(1, sheet.max_row):
            item = sheet.cell(row=count, column=1).value
            id = sheet.cell(row=count, column=9).value
            if id is None:
                print('Create a new category')
                cat_new = Category()
                cat_new.name = item
                cat_new.save()
            else:
                print('Create a new good')
